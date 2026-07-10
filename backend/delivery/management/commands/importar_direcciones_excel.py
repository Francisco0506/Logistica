from datetime import time
from pathlib import Path

import openpyxl
from django.core.management.base import BaseCommand

from delivery.models import Destino


class Command(BaseCommand):
    help = (
        "Importa destinos (direcciones, ventanas de recibo, días de entrega) desde "
        "el Excel de SAP a la tabla Destino. Uso:\n"
        "  python manage.py importar_direcciones_excel delivery/data/Carga_direcciones_LineNum_AdresType.xlsx"
    )

    def add_arguments(self, parser):
        parser.add_argument("ruta_excel", type=str)

    def handle(self, *args, **options):
        ruta = Path(options["ruta_excel"])
        if not ruta.exists():
            self.stderr.write(f"No existe el archivo: {ruta}")
            return

        wb = openpyxl.load_workbook(ruta, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]

        def to_time(v):
            if v is None:
                return None
            if isinstance(v, str) and ":" in v:
                h, m = v.split(":")
                return time(int(h), int(m))
            return None

        def es_si(v):
            return str(v).strip().upper() == "S" if v is not None else True

        creados, actualizados = 0, 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            data = dict(zip(headers, row))
            if not data.get("CardCode"):
                continue

            _, created = Destino.objects.update_or_create(
                card_code=str(data["CardCode"]),
                ship_to_code=str(data.get("Address") or f"Dir-{data.get('LineNum')}"),
                defaults={
                    "street": data.get("Street"),
                    "block": data.get("Block"),
                    "city": data.get("City"),
                    "zip_code": str(data.get("ZipCode") or ""),
                    "latitude": data.get("U_Latitud"),
                    "longitude": data.get("U_Longitud"),
                    "ini_recibo_1": to_time(data.get("U_IniRecibo1")),
                    "fin_recibo_1": to_time(data.get("U_FinRecibo1")),
                    "ent_lun": es_si(data.get("U_EntLun")),
                    "ent_mar": es_si(data.get("U_EntMar")),
                    "ent_mie": es_si(data.get("U_EntMie")),
                    "ent_jue": es_si(data.get("U_EntJue")),
                    "ent_vie": es_si(data.get("U_EntVie")),
                    "ent_sab": es_si(data.get("U_EntSab")),
                },
            )
            creados += 1 if created else 0
            actualizados += 0 if created else 1

        self.stdout.write(self.style.SUCCESS(
            f"Destinos: {creados} creados, {actualizados} actualizados."
        ))
